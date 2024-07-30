import openpyxl

wb = openpyxl.load_workbook("/mnt/sdcard/master.xlsx", data_only=True)
# wb = openpyxl.load_workbook("/mnt/sdcard/students.xlsx", data_only=True)

ws = wb.active
headers = [cell.value for cell in ws[1]]
print(headers)
data = []
# get file rows
for row in ws.iter_rows(min_row=2, values_only=True):
    data.append(
        [(row[i] if row[i] is not None else "") for i in range(len(headers) - 1)]
    )
nwb = openpyxl.Workbook()
nws = nwb.active

for cell, n in zip(["A1", "B1", "C1", "D1", "E1", "F1"], headers[:-1]):
    nws[cell] = n
for row in data:
    nws.append(row)

nws["A1"] = "name"

nwb.save("/mnt/sdcard/students.xlsx")
