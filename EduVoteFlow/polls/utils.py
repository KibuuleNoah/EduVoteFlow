from flask import abort, flash, url_for, redirect
from functools import wraps
from flask_login import current_user
from flask_wtf.csrf import os
from sqlalchemy import func
from EduVoteFlow.models import Candidate, Poll, School, Student, User, db
from openpyxl import load_workbook


def extract_excel_data(fileobject):
    """
    Coverts an Excel file object to dict

    :param fileobject: object to read the data from
    :returns : a list of rows in a dict form
    :rtype : list
    """
    wb = load_workbook(fileobject, data_only=True)
    ws = wb.active
    # get file cols
    headers = [cell.value for cell in ws[1]]
    data = []
    # get file rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_object = {
            headers[i]: (row[i] if row[i] is not None else "")
            for i in range(len(headers))
        }
        # add a row to the rows list data
        data.append(data_object)
    return data


def create_student_username(fullname: str, grade: str, section: str) -> str:
    """
    Generates a username for a student based on their full name, grade, and section.

    :param fullname: The student's full name.
    :type fullname: str
    :param grade: The student's grade.
    :type grade: str
    :param section: The student's section or stream.
    :type section: str
    :return: A string representing the student's username.
    :rtype: str
    """
    # Split the full name into parts
    subnames = fullname.strip().split(" ")
    firstname = subnames[0]

    # Use the first name if it has more than two characters
    if len(firstname) > 2:
        return f"{firstname}{grade}{section}".title()
    if len(subnames) >= 2:
        # Otherwise, use the second part of the name
        return f"{subnames[1]}{grade}{section}".title()
    raise NotImplemented("arg are not propery provided")


def create_student_password(grade, section, roll_no):
    """
    Generates a student's password based on their grade, section, and roll number.

    :param grade: The student's grade.
    :type grade: str
    :param section: The student's section or stream.
    :type section: str
    :param roll_no: The student's roll number.
    :type roll_no: str
    :return: A string representing the student's password.
    :rtype: str
    """

    return f"{grade}{section}{roll_no}"


# Check For Duplocate usernames and report them.
def flag_duplicate_usernames(data):
    """
    Checks for duplicate usernames in the given data and reports them.

    :param data: A list of dictionaries, each containing user information including 'username' and 'id'.
    :type data: list
    :return: A list of Student objects with duplicate usernames.
    :rtype: list
    """

    # code to get index of duplicate names
    def list_duplicates_of(seq, item):
        start_at = -1
        locs = []
        while True:
            try:
                loc = seq.index(item, start_at + 1)
            except ValueError:
                break
            else:
                locs.append(loc)
                start_at = loc
        return locs

    # get index of duplicates
    usernames = [data[x]["username"] for x in range(len(data))]
    all_indexes = [list_duplicates_of(usernames, u) for u in usernames]

    # removing duplicate indexes
    duplicateIndexes = []
    for d in all_indexes:
        if len(d) > 1:
            if d not in duplicateIndexes:
                duplicateIndexes.append(d)
    # get duplicate items from indexes
    duplicatedObjects = []
    for e in duplicateIndexes:
        for index in e:
            duplicatedObjects.append(data[index])

    # send back db objects
    student_objects = []
    for i in range(len(duplicatedObjects)):
        student = Student.query.filter_by(
            id=int(duplicatedObjects[i]["id"]), roll_no=duplicatedObjects[i]["roll_no"]
        ).first()
        student_objects.append(student)
    return student_objects


def created_student_user(student: Student, db):
    """
    Creates a new user entry for a student in the database.

    :param student: The student object for whom the user entry is being created.
    :type student: Student
    :param db: The database session object used to interact with the database.
    :type db: SQLAlchemy session
    :return: None
    """
    new_user = User(user_type="student", user_id=student.id)
    db.session.add(new_user)
    db.session.commit()


def active_sched_poll_required(func):
    """
    Decorator to ensure that only active and scheduled polls are accessible.

    :param func: The function to be decorated.
    :type func: function
    :return: The decorated function.
    :rtype: function
    """

    @wraps(func)
    def func_wrapper(poll_id):
        # Query the database to find the poll by school_id and poll_id
        poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()

        # If the poll is not found, return a 404 error
        if not poll:
            abort(404)

        # Check if the poll status is not "Archived"
        if poll.status != "Archived":
            return func(poll_id)
        else:
            # If the poll is archived, display a flash message and redirect to the polls home page
            flash(
                "This Poll has been Archived! The Dashboard cannot be viewed for an Archived Poll.",
                "danger",
            )
            return redirect(url_for("polls.polls_home"))

    return func_wrapper


# def active_sched_poll_required(func):
#     """
#     Active and Scheduled polls required
#     """
#
#     @wraps(func)
#     def func_wrapper(poll_id):
#         poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()
#         if not poll:
#             abort(404)
#         if poll.status != "Archived":
#             return func(poll_id)
#         else:
#             flash(
#                 f"This Poll has been Archived! The Dashboard cannot be viewed for a Archived Poll",
#                 "danger",
#             )
#             return redirect(url_for("polls.polls_home"))
#
#     return func_wrapper


def activepollrequired(func):
    """
    Decorator to ensure that only active polls are accessible.

    :param func: The function to be decorated.
    :type func: function
    :return: The decorated function.
    :rtype: function
    """

    @wraps(func)
    def func_wrapper(poll_id, *args):
        # Query the database to find the poll by school_id and poll_id
        poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()

        # If the poll is not found, return a 404 error
        if not poll:
            abort(404)

        # Check if the poll status is "Active"
        if poll.status == "Active":
            return func(poll_id, *args)
        elif poll.status == "Scheduled":
            # If the poll is scheduled, display a flash message and redirect to the dashboard home page
            flash(
                "This Poll hasn't Started, This Action Can't be done on Scheduled Polls",
                "danger",
            )
            return redirect(url_for("polls.dashboard_home", poll_id=poll_id))
        else:
            # If the poll is archived or completed, display a flash message and redirect to the polls home page
            flash(
                f"This Poll has been {poll.status}! It can't make the task",
                "danger",
            )
            return redirect(url_for("polls.polls_home"))

    return func_wrapper


# def activepollrequired(func):
#     """
#     Active polls required
#     """
#
#     @wraps(func)
#     def func_wrapper(poll_id, *args):
#         poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()
#         if not poll:
#             abort(404)
#         if poll.status == "Active":
#             return func(poll_id, *args)
#         elif poll.status == "Scheduled":
#             flash(
#                 "This Poll hasn't Started, This Action Can't be done on Scheduled Polls",
#                 "danger",
#             )
#             return redirect(url_for("polls.dashboard_home", poll_id=poll_id))
#         else:
#             flash(
#                 f"This Poll has been {poll.status}! It can't make the task",
#                 "danger",
#             )
#             return redirect(url_for("polls.polls_home"))
#
#     return func_wrapper


def scheduledpollrequired(func):
    """
    Decorator to ensure that only scheduled polls are accessible.

    :param func: The function to be decorated.
    :type func: function
    :return: The decorated function.
    :rtype: function
    """

    @wraps(func)
    def func_wrapper(poll_id, *args):
        # Query the database to find the poll by school_id and poll_id
        poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()

        # If the poll is not found, return a 404 error
        if not poll:
            abort(404)

        # Check if the poll status is "Scheduled"
        if poll.status == "Scheduled":
            return func(poll_id, *args)
        elif poll.status == "Active":
            # If the poll is active, display a flash message and redirect to the dashboard home page
            flash("This Poll is Already Active, This Action Can't be done", "danger")
            return redirect(url_for("polls.dashboard_home", poll_id=poll_id))
        else:
            # If the poll is archived or completed, display a flash message and redirect to the polls home page
            flash(f"This Poll has been {poll.status}! It can't make the task", "danger")
            return redirect(url_for("polls.polls_home"))

    return func_wrapper


# def scheduledpollrequired(func):
#     """
#     Scheduled polls required
#     """
#
#     @wraps(func)
#     def func_wrapper(poll_id, *args):
#         poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()
#         if not poll:
#             abort(404)
#         if poll.status == "Scheduled":
#             return func(poll_id, *args)
#         elif poll.status == "Active":
#             flash("This Poll is Already Active,This Action Can't be done", "danger")
#             return redirect(url_for("polls.dashboard_home", poll_id=poll_id))
#         else:
#             flash(
#                 f"This Poll has been {poll.status}! It can't make the task",
#                 "danger",
#             )
#             return redirect(url_for("polls.polls_home"))
#
#     return func_wrapper


def create_candidate_record(candidate: dict, school_id: int, poll: Poll):
    """
    Creates and returns a list of Candidate objects based on the provided candidate information.

    :param candidate: A dictionary containing the candidate's details such as name, house, grade, gender, post, and slogan.
    :type candidate: dict
    :param school_id: The ID of the school where the poll is taking place.
    :type school_id: int
    :param poll: The Poll object associated with the candidate.
    :type poll: Poll
    :return: A list containing the newly created Candidate object.
    :rtype: list
    """
    candidate_objects = []

    # Create a new Candidate object using the provided details
    new_candidate = Candidate(
        school_id=school_id,
        poll_id=poll.id,
        full_name=candidate["student_name"].title(),  # Capitalize the full name
        house=candidate.get(
            "house", ""
        ).title(),  # Capitalize the house name, default to empty string if not provided
        grade=candidate["grade"],
        gender=candidate["gender"],
        post=candidate["post"].title(),  # Capitalize the post name
        slogan=candidate["slogan"],
    )

    # Add the new Candidate object to the list
    candidate_objects.append(new_candidate)

    return candidate_objects


def get_poll_link(school_id: int, poll_id: int):
    """
    Generates the URL for the splash screen of a specific poll.

    :param school_id: The ID of the school where the poll is taking place.
    :type school_id: int
    :param poll_id: The ID of the poll.
    :type poll_id: int
    :return: The full URL for the poll's splash screen.
    :rtype: str
    """
    # Query the database to find the poll by poll_id
    poll = Poll.query.filter_by(id=poll_id).first()

    # Generate and return the URL for the splash screen of the specified poll
    return url_for(
        "election.splash_screen",  # Endpoint for the splash screen view
        school_id=school_id,  # Pass the school ID as a parameter
        poll_id=poll.id,  # Pass the poll ID as a parameter
        _external=True,  # Generate an absolute URL
    )


def get_unopposed_candidates(poll_id: int, school_id: int):
    """
    Retrieves candidates who are unopposed in a specific poll.

    :param poll_id: The ID of the poll.
    :type poll_id: int
    :param school_id: The ID of the school where the poll is taking place.
    :type school_id: int
    :return: A query object containing candidates who have no opponents for their respective posts.
    :rtype: sqlalchemy.orm.query.Query
    """
    # Query the database to find candidates who are unopposed
    unopposed_candidates = (
        Candidate.query.filter_by(poll_id=poll_id, school_id=school_id)
        .group_by(Candidate.post)  # Group by the candidate's post
        .having(
            func.count(Candidate.id) == 1
        )  # Select groups with exactly one candidate
    )

    return unopposed_candidates


def get_opposed_candidates(poll_id: int, school_id: int):
    """
    Retrieves candidates who are opposed (i.e., have competitors) in a specific poll.

    :param poll_id: The ID of the poll.
    :type poll_id: int
    :param school_id: The ID of the school where the poll is taking place.
    :type school_id: int
    :return: A query object containing candidates who have opponents for their respective posts.
    :rtype: sqlalchemy.orm.query.Query
    """
    # Subquery to find posts with more than one candidate
    subquery = (
        db.session.query(Candidate.post)
        .group_by(Candidate.post)  # Group by the candidate's post
        .having(
            func.count(Candidate.id) > 1
        )  # Select posts with more than one candidate
        .subquery()
    )

    # Main query to get all candidates for posts with more than one candidate
    opposed_candidates = Candidate.query.filter(Candidate.post.in_(subquery)).filter_by(
        poll_id=poll_id, school_id=school_id  # Filter by poll ID  # Filter by school ID
    )

    return opposed_candidates


def save_candidates(
    school: School, poll: Poll, posts: list[str], candidates: list[dict], db
):
    """
    Saves candidate records to the database.

    :param school: The School object where the candidates are running.
    :type school: School
    :param poll: The Poll object associated with the candidates.
    :type poll: Poll
    :param posts: A list of valid post names for the poll.
    :type posts: list[str]
    :param candidates: A list of dictionaries, each containing candidate details such as student_name, house, gender, post, and slogan.
    :type candidates: list[dict]
    :param db: The database session object used to interact with the database.
    :type db: SQLAlchemy session
    :return: True if candidates were successfully saved, False if there are missing required columns.
    :rtype: bool
    """
    # Define the expected columns for candidate dictionaries
    expected_cols = {"student_name", "house", "gender", "post", "slogan"}
    # Get the actual columns from the first candidate dictionary
    actual_cols = set(candidates[0].keys())

    # Check if any expected columns are missing in the candidate dictionaries
    if expected_cols - actual_cols:
        return False

    # Iterate over each candidate dictionary
    for candidate in candidates:
        # Only process candidates with valid posts
        if candidate["post"] in posts:
            # Save the candidate record to the database
            db.session.bulk_save_objects(
                create_candidate_record(candidate, school.id, poll)
            )
            db.session.commit()

    return True


def get_schpolldir_path(school: School, poll: Poll, sch_path: bool = False) -> str:
    """
    Generates the directory path for storing data related to a specific school and poll.

    :param school: The School object for which the directory path is generated.
    :type school: School
    :param poll: The Poll object related to the directory path.
    :type poll: Poll
    :param sch_path: Boolean flag indicating whether to return the path for the school directory only. If False, returns the path including the poll directory.
    :type sch_path: bool
    :return: The full directory path as a string.
    :rtype: str
    """
    # Define the directory names based on school and poll details
    school_dir = f"{school.abbr}{school.id}"
    poll_dir = f"{poll.id}"

    # Generate the path based on whether the school path or full path is requested
    if sch_path:
        path = f"{os.getcwd()}/EduVoteFlow/static/DataStore/{school_dir}"
    else:
        path = f"{os.getcwd()}/EduVoteFlow/static/DataStore/{school_dir}/{poll_dir}"

    return path


def posttotal_votecasts(candidates: list) -> int:
    """
    Calculates the total number of votes cast for a post given the list of candidates.

    :param candidates: A list of dictionaries, each representing a candidate with a "votes" key.
    :type candidates: list
    :return: The total number of votes cast for all candidates.
    :rtype: int
    """
    total = 0

    # Iterate over each candidate and sum their votes
    for candidate in candidates:
        total += candidate["votes"]

    return total


def candidates_percent(votes_casted: int, candidate: dict) -> dict:
    """
    Calculates the percentage of votes a candidate has received out of the total votes cast at a pot.

    :param votes_casted: The total number of votes cast across all candidates at that post.
    :type votes_casted: int
    :param candidate: A dictionary representing the candidate, with a "votes" key indicating the number of votes the candidate received.
    :type candidate: dict
    :return: The updated candidate dictionary with an additional "percentage" key showing the vote percentage.
    :rtype: dict
    """
    candidate_votes = candidate.get(
        "votes", 0
    )  # Get the number of votes for the candidate
    percentage = 0

    # Calculate the percentage if there are votes casted and the candidate has received votes
    if votes_casted > 0:
        percentage = round((candidate_votes / votes_casted) * 100, 1)

    # Add the percentage to the candidate dictionary
    candidate["percentage"] = percentage

    return candidate


def rank_candidates_by_post(post_candidates: dict) -> dict:
    """
    Ranks candidates by their votes for each post and categorizes them as winners or losers.

    :param post_candidates: A dictionary where keys are post names and values are lists of candidate dictionaries.
        Each candidate dictionary contains candidate details including the number of votes received.
    :type post_candidates: dict
    :return: A dictionary with overall winners, losers, and detailed results per post.
    :rtype: dict
    """
    # Define a lambda function to sort candidates by their votes
    sort_by_votes = lambda x: x["votes"]

    # Initialize result containers
    overall_winners = []
    overall_losers = []
    detailed_results = {}

    # Process each post and its candidates
    for post, candidates in post_candidates.items():
        # Sort candidates by number of votes in descending order
        sorted_candidates = sorted(candidates, key=sort_by_votes, reverse=True)
        votes_casted = posttotal_votecasts(sorted_candidates)

        # Add percentage of votes to each candidate
        sorted_candidates = [
            candidates_percent(votes_casted, candidate)
            for candidate in sorted_candidates
        ]

        # Initialize results for the current post
        if post not in detailed_results:
            detailed_results[post] = {
                "votes_casted": votes_casted,
                "winner_votes": 0,
                "winners": [],
                "losers": [],
                "any_tie": False,
                "any_voted": True,
            }

        post_results = detailed_results[post]

        # Determine winners and losers
        for idx, candidate in enumerate(sorted_candidates):
            # If all candidates have zero votes
            if idx == 0 and candidate["votes"] == 0:
                post_results["any_voted"] = False
                post_results["losers"].extend(sorted_candidates)
                overall_losers.extend(sorted_candidates)
                break
            # If multiple candidates have the same maximum votes
            elif (
                post_results["winners"]
                and candidate["votes"] == post_results["winner_votes"]
            ):
                post_results["winners"].append(candidate)
                if not post_results["any_tie"]:
                    post_results["any_tie"] = True
            # If the current candidate has fewer votes than the current winners
            elif (
                post_results["winners"]
                and candidate["votes"] < post_results["winner_votes"]
            ):
                post_results["losers"].extend(sorted_candidates[idx:])
                overall_losers.extend(sorted_candidates[idx:])
                break
            else:
                post_results["winners"].append(candidate)
                if post_results["winner_votes"] == 0:
                    post_results["winner_votes"] = candidate["votes"]

        # Collect winners if there are no ties
        if not post_results["any_tie"]:
            overall_winners.extend(post_results["winners"])

    return {
        "winners": overall_winners,
        "losers": overall_losers,
        "res_obj": detailed_results,
    }


#
# sort_by_votes = lambda x: x["votes"]
# def candidates_rank(post_candidates: dict):
#     res = {}
#     winners = []
#     losers = []
#     for post in post_candidates.keys():
#         sorted_candidates = sorted(post_candidates[post], key=sort_by_votes, reverse=True)
#         votes_casted = posttotal_votecasts(sorted_candidates)
#         sorted_candidates = [
#             candidates_percent(votes_casted, c) for c in sorted_candidates
#         ]
#         if post not in res:
#             res[post] = {
#                 "votes_casted": votes_casted,
#                 "winner_votes": 0,
#                 "winners": [],
#                 "losers": [],
#                 "any_tie": False,
#                 "any_voted": True,
#             }
#         res_post = res[post]
#         for idx, cand in enumerate(sorted_candidates):
#             # if all candidates have zero votes
#             if idx == 0 and cand["votes"] == 0:
#                 res[post]["any_voted"] = False
#                 res[post]["losers"].extend(sorted_candidates)
#                 losers.extend(sorted_candidates)
#                 break
#             # if muiltple candidates have the same max votes
#             elif res_post["winners"] and cand["votes"] == res_post["winner_votes"]:
#                 res[post]["winners"].append(cand)
#                 if not res_post["any_tie"]:
#                     res[post]["any_tie"] = True
#             # if thr current candidate is already a loser
#             elif res_post["winners"] and cand["votes"] < res_post["winner_votes"]:
#                 res[post]["losers"].extend(sorted_candidates[idx:])
#                 losers.extend(sorted_candidates[idx:])
#                 break
#             else:
#                 res[post]["winners"].append(cand)
#                 if res_post["winner_votes"] == 0:
#                     res[post]["winner_votes"] = cand["votes"]
#         if not res[post]["any_tie"]:
#             winners.extend(res[post]["winners"])
#     return {"winners": winners, "losers": losers, "res_obj": res}


def get_gender(gender_indicator: str) -> str:
    """
    Determines the gender based on the provided string indicator.

    :param gender_indicator: A string where the first character indicates gender ('M' or 'F').
    :type gender_indicator: str
    :return: "Male" if the indicator starts with 'M' or 'm', otherwise "Female".
    :rtype: str
    """
    if gender_indicator and gender_indicator[0].lower() == "m":
        return "Male"
    return "Female"


def validate_student_cols(students: list[dict]) -> bool:
    """
    Validates if the provided list of student dictionaries contains all the required columns.

    :param students: A list of dictionaries, each representing a student with potential columns.
    :type students: list[dict]
    :return: True if all student dictionaries have the required columns, otherwise False.
    :rtype: bool
    """
    if not students:
        return False

    # Define the expected columns for student dictionaries
    expected_cols = {"student_name", "grade", "gender", "section", "roll_no"}
    # Get the actual columns from the first student dictionary
    actual_cols = set(students[0].keys())

    # Check if any expected columns are missing
    if expected_cols - actual_cols:
        return False

    return True
