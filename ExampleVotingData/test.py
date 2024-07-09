
from openpyxl import load_workbook

def extract_excel_data(fileobject):
    wb = load_workbook(fileobject, data_only=True)
    ws = wb['Sheet1']
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_object = {headers[i]: (row[i] if row[i] is not None else "") for i in range(len(headers))}
        data.append(data_object)
    return data
