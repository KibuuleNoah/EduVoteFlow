# import openpyxl
#
# wb = openpyxl.load_workbook("/mnt/sdcard/students.xlsx")
# # wb = openpyxl.load_workbook("/mnt/sdcard/candidates.xlsx")
# # ws = wb.active
# #
# # print([c.value for c in ws[1]])
# # ['student_name', 'grade', 'section', 'gender', 'house', 'roll_no']
# # ['student_name', 'house', 'grade', 'gender', 'slogan', 'post']
#
#
# # Create a new Excel file
# wb = openpyxl.Workbook()
#
# # Select the first sheet
# sheet = wb.active
#
# # Set the header row
# sheet["A1"] = "name"
# sheet["B1"] = "house"
# sheet["C1"] = "grade"
# sheet["D1"] = "gender"
# sheet["E1"] = "slogan"
# sheet["F1"] = "post"
#
#
# # Add data to the sheet using tuples
# data = [
#     ("John Doe", "H1", 11, "Male", "Hello World!", "P1"),
#     ("Jane Smith", "H2", 7, "Female", "Goodbye World!", "P1"),
#     ("Bob Johnson", "H3", 8, "Male", "Hello Again!", "P1"),
#     ("Noah Johnson", "H1", 8, "Male", "Hello Again and !", "P2"),
#     ("Tech Tim", "H3", 8, "female", "Hello Again noomm!", "P2"),
#     ("Tristar Mosh", "H3", 8, "Male", "Hello Again noomm!", "P3"),
#     ("Tech Mosh", "H2", 8, "Male", "Hello Again noomm!", "P3"),
#     ("Moxie Tim", "H3", 8, "Female", "Hello Tick!", "P3"),
#     ("John Felt", "H1", 8, "Male", "Hello Again noomm!", "P3"),
#     ("John Mosh", "H3", 11, "Male", "Hello Boom!", "P4"),
# ]
#
# for row in data:
#     sheet.append(row)
#
# # Save the file
# wb.save("/mnt/sdcard/candidates.xlsx")
# # wb.s
# # av()
#
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Function to create an Excel file
def create_excel_file(file_name, headers, data):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)  # Append headers
    for row in data:
        ws.append(row)  # Append each row of data
    wb.save(file_name)


# Function to generate HTML table from data
def generate_html_table(headers, data):
    table = '<table border="1">\n'
    table += "  <tr>\n"
    for header in headers:
        table += f"    <th>{header}</th>\n"
    table += "  </tr>\n"
    for row in data:
        table += "  <tr>\n"
        for item in row:
            table += f"    <td>{item}</td>\n"
        table += "  </tr>\n"
    table += "</table>"
    return table


# Data for the first Excel file
headers1 = ["name", "grade", "section", "gender", "house", "roll_no"]
data1 = [
    ["Alice Mary", "A", "1", "Female", "Red", 1],
    ["Bob Mosh", "B", "2", "Male", "Blue", 2],
    ["Charlie Tech", "C", "1", "Male", "Green", 3],
    ["Diana", "A", "2", "Female", "Yellow", 4],
    ["Eve Mia", "B", "3", "Female", "Red", 5],
]

# Data for the second Excel file
headers2 = ["name", "house", "grade", "gender", "slogan", "post"]
data2 = [
    ["Alice Jane", "Red", "A", "Female", "Brave and Bold", "Captain"],
    ["Mosh Bob", "Blue", "B", "Male", "Strong and True", "Vice-Captain"],
    ["Charlie Puth", "Green", "C", "Male", "Wise and Just", "Member"],
    ["Diana Omos", "Yellow", "A", "Female", "Bright and Cheery", "Member"],
    ["Eve Adam", "Red", "B", "Female", "Fearless and Free", "Member"],
]

# Create Excel files
create_excel_file("students1.xlsx", headers1, data1)
create_excel_file("students2.xlsx", headers2, data2)

# Generate HTML tables
html_table1 = generate_html_table(headers1, data1)
html_table2 = generate_html_table(headers2, data2)

# Save HTML tables to files
with open("students1.html", "w") as f:
    f.write(html_table1)

with open("students2.html", "w") as f:
    f.write(html_table2)

print("Excel files and HTML tables created successfully.")
