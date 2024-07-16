from flask import abort, flash, url_for, redirect
from functools import wraps
from flask_login import current_user
from sqlalchemy import func
from VoteFlow.models import Candidate, Poll, School, Student, User
from openpyxl import load_workbook


def extract_excel_data(fileobject):
    wb = load_workbook(fileobject, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data_object = {
            headers[i]: (row[i] if row[i] is not None else "")
            for i in range(len(headers))
        }
        data.append(data_object)
    print(headers)
    print(data)
    return data


def created_student_username(fullname, grade, section):
    subnames = fullname.strip().split(" ")
    firstname = subnames[0]
    # check for single letter first name
    if len(firstname) > 2:
        return f"{firstname}{grade}{section}"
    return f"{subnames[1]}{grade}{section}"


def created_student_password(grade, section, roll_no):
    return f"{grade}{section}{roll_no}"


# Check For Duplocate usernames and report them.
def flag_duplicate_usernames(data):
    # code to get index of duplicate names
    def list_duplicates_of(seq, item):
        start_at = -1
        locs = []
        while True:
            try:
                loc = seq.index(item, start_at + 1)
            except ValueError:
                break
            else:
                locs.append(loc)
                start_at = loc
        return locs

    # get index of duplicates
    usernames = [data[x]["username"] for x in range(len(data))]
    all_indexes = [list_duplicates_of(usernames, u) for u in usernames]

    # removing duplicate indexes
    duplicateIndexes = []
    for d in all_indexes:
        if len(d) > 1:
            if d not in duplicateIndexes:
                duplicateIndexes.append(d)
    # get duplicate items from indexes
    duplicatedObjects = []
    for e in duplicateIndexes:
        for index in e:
            duplicatedObjects.append(data[index])

    # send back db objects
    student_objects = []
    for i in range(len(duplicatedObjects)):
        student = Student.query.filter_by(
            id=int(duplicatedObjects[i]["id"]), roll_no=duplicatedObjects[i]["roll_no"]
        ).first()
        student_objects.append(student)
    return student_objects


def created_student_user(student: Student, db):
    new_user = User(user_type="student", user_id=student.id)
    db.session.add(new_user)
    db.session.commit()


def active_sched_poll_required(func):
    """
    Active and Scheduled polls required
    """

    @wraps(func)
    def func_wrapper(poll_id):
        poll = Poll.query.filter_by(school_id=current_user.id, id=poll_id).first()
        if not poll:
            abort(404)
        if poll.status != "Archived":
            return func(poll_id)
        else:
            flash(
                f"This Poll has been Archived! The Dashboard cannot be viewed for a Archived Poll",
                "danger",
            )
            return redirect(url_for("polls.home"))

    return func_wrapper


def add_record(candidate: dict, school_id: int, poll: Poll):
    candidate_objects = []
    new_candidate = Candidate(
        school_id=school_id,
        poll_id=poll.id,
        full_name=candidate["student_name"],
        house="",
        gender="",
        post=f"[{candidate['post']}]",
        logo=f"{candidate['student_name'].replace(' ', '_')}.jpg",
        slogan=candidate["slogan"],
        votes=0,
    )
    candidate_objects.append(new_candidate)
    return candidate_objects


def get_poll_link(school_id: int, poll_id: int):
    poll = Poll.query.filter_by(id=poll_id).first()
    return url_for(
        "election.splash_screen", school_id=school_id, poll_id=poll.id, _external=True
    )


def get_unopposed_candidates(db):
    unopposed_candidates = (
        db.session.query(Candidate)
        .group_by(Candidate.post)
        .having(func.count(Candidate.id) == 1)
        .all()
    )

    return unopposed_candidates


def save_candidates(
    school: School, poll: Poll, posts: list[str], candidates: list[dict], db
):
    for post in posts:
        for candidate in candidates:
            if post == candidate["post"]:
                db.session.bulk_save_objects(add_record(candidate, school.id, poll))
                db.session.commit()
        print("-" * 30)
